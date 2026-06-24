from mcp.server.fastmcp import FastMCP
import httpx
import json
import os
import re
import difflib
from datetime import datetime
from pathlib import Path

mcp = FastMCP("my-server")

WIKI_API = "https://en.wikipedia.org/api/rest_v1"
HEADERS = {"User-Agent": "MCPServer/1.0 (https://github.com/example; contact@example.com)"}

# ── Basic ──────────────────────────────────────────────

@mcp.resource("greeting://{name}")
def get_greeting(name: str) -> str:
    return f"Hello, {name}!"

@mcp.tool()
def add(a: int, b: int) -> int:
    """Add two numbers."""
    return a + b

# ═══════════════════════════════════════════════════════
#  EXCEL (openpyxl)
# ═══════════════════════════════════════════════════════

def _get_excel(path: str):
    from openpyxl import load_workbook, Workbook
    p = Path(path).expanduser().resolve()
    if not p.exists():
        wb = Workbook()
        wb.save(p)
        return wb, p
    return load_workbook(p), p

@mcp.tool()
def excel_read(path: str, sheet: str = "", cell: str = "") -> str:
    """Read an Excel file. If no sheet given, list sheets. If no cell, show sheet contents.
    Path can be relative or absolute. Example: excel_read data.xlsx Sheet1"""
    wb, p = _get_excel(path)
    try:
        if not sheet:
            return f"Sheets: {', '.join(wb.sheetnames)}"
        ws = wb[sheet]
        if cell:
            return f"{cell}: {ws[cell].value}"
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append("\t".join(str(c) if c is not None else "" for c in row))
        return f"File: {p.name} / Sheet: {sheet}\n" + "\n".join(rows[:50])
    finally:
        wb.close()

@mcp.tool()
def excel_write(path: str, sheet: str, cell: str, value: str) -> str:
    """Write a value to an Excel cell. Creates file/sheet if needed.
    Example: excel_write data.xlsx Sheet1 A1 hello"""
    wb, p = _get_excel(path)
    try:
        if sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            ws = wb.create_sheet(sheet)
        ws[cell] = value
        wb.save(p)
        return f"Written '{value}' to {p.name}[{sheet}]!{cell}"
    finally:
        wb.close()

@mcp.tool()
def excel_create(path: str, headers: str = "") -> str:
    """Create a new Excel file. Optionally add comma-separated headers.
    Example: excel_create data.xlsx Name,Email,Age"""
    p = Path(path).expanduser().resolve()
    from openpyxl import Workbook
    wb = Workbook()
    if headers:
        ws = wb.active
        ws.title = "Sheet1"
        for col, h in enumerate(headers.split(","), 1):
            ws.cell(row=1, column=col, value=h.strip())
    wb.save(p)
    return f"Created {p}"

@mcp.tool()
def excel_add_row(path: str, sheet: str, values: str) -> str:
    """Append a row to an Excel sheet. Values are comma-separated.
    Example: excel_add_row data.xlsx Sheet1 Alice,alice@x.com,30"""
    wb, p = _get_excel(path)
    try:
        if sheet not in wb.sheetnames:
            return f"Sheet '{sheet}' not found"
        ws = wb[sheet]
        row = ws.max_row + 1
        for col, v in enumerate(values.split(","), 1):
            ws.cell(row=row, column=col, value=v.strip())
        wb.save(p)
        return f"Row added to {p.name}[{sheet}]"
    finally:
        wb.close()

# ═══════════════════════════════════════════════════════
#  GMAIL (requires Google Cloud OAuth setup)
# ═══════════════════════════════════════════════════════

GMAIL_CREDS = os.getenv("GMAIL_CREDS_PATH", "")

def _get_gmail_service():
    if not GMAIL_CREDS or not Path(GMAIL_CREDS).exists():
        return None
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build

    creds = None
    token_path = Path(GMAIL_CREDS).parent / "gmail_token.json"
    if token_path.exists():
        creds = Credentials.from_authorized_user_file(str(token_path), ["https://www.googleapis.com/auth/gmail.modify"])
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(GMAIL_CREDS, ["https://www.googleapis.com/auth/gmail.modify"])
            creds = flow.run_local_server(port=0)
        with open(str(token_path), "w") as f:
            f.write(creds.to_json())
    return build("gmail", "v1", credentials=creds)

@mcp.tool()
def gmail_setup(creds_path: str) -> str:
    """Set up Gmail by pointing to your Google OAuth client_secret.json.
    Get it from https://console.cloud.google.com -> APIs & Services -> Credentials
    Set GMAIL_CREDS_PATH env var or pass path here. Used once to authenticate."""
    return (f"Set GMAIL_CREDS_PATH={creds_path} as environment variable and "
            f"restart the server. On first call, it will open a browser for OAuth.")

@mcp.tool()
def gmail_send(to: str, subject: str, body: str) -> str:
    """Send an email via Gmail. Requires Google Cloud OAuth setup first.
    Example: gmail_send friend@x.com Hello Check this out"""
    service = _get_gmail_service()
    if not service:
        return "Gmail not configured. Call gmail_setup first, or set GMAIL_CREDS_PATH env var."
    from email.mime.text import MIMEText
    import base64
    msg = MIMEText(body)
    msg["To"] = to
    msg["Subject"] = subject
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    service.users().messages().send(userId="me", body={"raw": raw}).execute()
    return f"Email sent to {to}"

@mcp.tool()
def gmail_inbox(max_results: int = 10) -> str:
    """List recent emails from Gmail inbox."""
    service = _get_gmail_service()
    if not service:
        return "Gmail not configured. Call gmail_setup first."
    results = service.users().messages().list(userId="me", maxResults=max_results, q="in:inbox").execute()
    msgs = results.get("messages", [])
    lines = []
    for m in msgs:
        msg = service.users().messages().get(userId="me", id=m["id"], format="metadata",
                                             metadataHeaders=["From", "Subject", "Date"]).execute()
        headers = {h["name"]: h["value"] for h in msg["payload"]["headers"]}
        lines.append(f"From: {headers.get('From', '?')}")
        lines.append(f"Subj: {headers.get('Subject', '?')}")
        lines.append(f"Date: {headers.get('Date', '?')}")
        lines.append("")
    return "\n".join(lines) if lines else "No emails found."

@mcp.tool()
def gmail_search(query: str, max_results: int = 10) -> str:
    """Search Gmail. Example: gmail_search from:alice"""
    service = _get_gmail_service()
    if not service:
        return "Gmail not configured."
    results = service.users().messages().list(userId="me", maxResults=max_results, q=query).execute()
    msgs = results.get("messages", [])
    lines = [f"Found {len(msgs)} email(s):"]
    for m in msgs:
        msg = service.users().messages().get(userId="me", id=m["id"], format="metadata",
                                             metadataHeaders=["From", "Subject"]).execute()
        headers = {h["name"]: h["value"] for h in msg["payload"]["headers"]}
        lines.append(f"  {headers.get('From', '?')} | {headers.get('Subject', '?')}")
    return "\n".join(lines) if msgs else "No results."

# ═══════════════════════════════════════════════════════
#  GOOGLE DOCS (requires Google Cloud OAuth setup)
# ═══════════════════════════════════════════════════════

DOCS_CREDS = os.getenv("DOCS_CREDS_PATH", "")

def _get_docs_service():
    if not DOCS_CREDS or not Path(DOCS_CREDS).exists():
        return None
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build

    creds = None
    token_path = Path(DOCS_CREDS).parent / "docs_token.json"
    if token_path.exists():
        creds = Credentials.from_authorized_user_file(str(token_path),
            ["https://www.googleapis.com/auth/documents"])
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(DOCS_CREDS,
                ["https://www.googleapis.com/auth/documents"])
            creds = flow.run_local_server(port=0)
        with open(str(token_path), "w") as f:
            f.write(creds.to_json())
    return build("docs", "v1", credentials=creds)

@mcp.tool()
def docs_setup(creds_path: str) -> str:
    """Set up Google Docs by pointing to your OAuth client_secret.json.
    Set DOCS_CREDS_PATH env var or pass path here."""
    return (f"Set DOCS_CREDS_PATH={creds_path} as environment variable and "
            f"restart the server.")

@mcp.tool()
def docs_create(title: str, content: str = "") -> str:
    """Create a new Google Doc. Optionally with initial content.
    Example: docs_create MyDoc Hello, world!"""
    service = _get_docs_service()
    if not service:
        return "Google Docs not configured. Call docs_setup first."
    doc = service.documents().create(body={"title": title}).execute()
    doc_id = doc["documentId"]
    if content:
        service.documents().batchUpdate(
            documentId=doc_id,
            body={"requests": [{
                "insertText": {"location": {"index": 1}, "text": content}
            }]}
        ).execute()
    return f"Created: {title}\nURL: https://docs.google.com/document/d/{doc_id}"

@mcp.tool()
def docs_read(doc_url_or_id: str) -> str:
    """Read a Google Doc by URL or document ID.
    Example: docs_read https://docs.google.com/document/d/ABC123"""
    doc_id = doc_url_or_id.split("/d/")[-1].split("/")[0] if "/" in doc_url_or_id else doc_url_or_id
    service = _get_docs_service()
    if not service:
        return "Google Docs not configured."
    doc = service.documents().get(documentId=doc_id).execute()
    title = doc.get("title", "")
    body = doc.get("body", {}).get("content", [])
    text = ""
    for el in body:
        if "paragraph" in el:
            for run in el["paragraph"].get("elements", []):
                if "textRun" in run:
                    text += run["textRun"].get("content", "")
    return f"Title: {title}\n\n{text[:5000]}"

@mcp.tool()
def docs_append(doc_url_or_id: str, text: str) -> str:
    """Append text to an existing Google Doc."""
    doc_id = doc_url_or_id.split("/d/")[-1].split("/")[0] if "/" in doc_url_or_id else doc_url_or_id
    service = _get_docs_service()
    if not service:
        return "Google Docs not configured."
    doc = service.documents().get(documentId=doc_id).execute()
    end_index = doc.get("body", {}).get("content", [{}])[-1].get("endIndex", 1)
    service.documents().batchUpdate(
        documentId=doc_id,
        body={"requests": [{"insertText": {"location": {"index": end_index - 1}, "text": text}}]}
    ).execute()
    return f"Text appended to document."

# ═══════════════════════════════════════════════════════
#  UNIT CONVERTER
# ═══════════════════════════════════════════════════════

CONVERSIONS = {
    ("km", "mi"): 0.621371, ("mi", "km"): 1.60934,
    ("m", "ft"): 3.28084, ("ft", "m"): 0.3048,
    ("cm", "in"): 0.393701, ("in", "cm"): 2.54,
    ("kg", "lb"): 2.20462, ("lb", "kg"): 0.453592,
    ("g", "oz"): 0.035274, ("oz", "g"): 28.3495,
    ("l", "gal"): 0.264172, ("gal", "l"): 3.78541,
    ("ml", "fl_oz"): 0.033814, ("fl_oz", "ml"): 29.5735,
    ("c", "f"): lambda c: c * 9/5 + 32,
    ("f", "c"): lambda f: (f - 32) * 5/9,
    ("c", "k"): lambda c: c + 273.15,
    ("k", "c"): lambda k: k - 273.15,
    ("km_h", "mph"): 0.621371, ("mph", "km_h"): 1.60934,
    ("m_s", "km_h"): 3.6, ("km_h", "m_s"): 1/3.6,
}

UNITS = {
    "km": "Kilometer", "mi": "Mile", "m": "Meter", "ft": "Foot",
    "cm": "Centimeter", "in": "Inch", "kg": "Kilogram", "lb": "Pound",
    "g": "Gram", "oz": "Ounce", "l": "Liter", "gal": "Gallon",
    "ml": "Milliliter", "fl_oz": "Fluid Ounce", "c": "Celsius",
    "f": "Fahrenheit", "k": "Kelvin", "km_h": "km/h", "mph": "mph",
    "m_s": "m/s",
}

@mcp.tool()
def convert(value: float, from_unit: str, to_unit: str) -> str:
    """Convert between units. Length: km,mi,m,ft,cm,in. Weight: kg,lb,g,oz.
    Volume: l,gal,ml,fl_oz. Temp: c,f,k. Speed: km_h,mph,m_s."""
    key = (from_unit.lower(), to_unit.lower())
    if key not in CONVERSIONS:
        return f"Unknown: {from_unit} -> {to_unit}"
    factor = CONVERSIONS[key]
    result = factor(value) if callable(factor) else value * factor
    return f"{value} {UNITS.get(from_unit, from_unit)} = {result:.4f} {UNITS.get(to_unit, to_unit)}"

@mcp.tool()
def units() -> str:
    """List available units."""
    return """Length:  km, mi, m, ft, cm, in
Weight:  kg, lb, g, oz
Volume:  l, gal, ml, fl_oz
Temp:    c, f, k
Speed:   km_h, mph, m_s
Usage: convert 100 km mi"""

# ═══════════════════════════════════════════════════════
#  CURRENCY
# ═══════════════════════════════════════════════════════

@mcp.tool()
def currency(amount: float, from_curr: str, to_curr: str) -> str:
    """Convert currency. Example: currency 100 usd inr"""
    try:
        resp = httpx.get(f"https://api.exchangerate-api.com/v4/latest/{from_curr.upper()}", timeout=10)
        data = resp.json()
        rate = data["rates"].get(to_curr.upper())
        if not rate:
            return f"Unknown: {to_curr}"
        return f"{amount} {from_curr.upper()} = {amount * rate:.2f} {to_curr.upper()}\nRate: 1 {from_curr.upper()} = {rate} {to_curr.upper()}"
    except Exception as e:
        return f"Error: {e}"

@mcp.tool()
def currencies() -> str:
    """List common currency codes."""
    return "USD EUR GBP JPY CNY INR AUD CAD CHF KRW BRL MXN SGD HKD SEK NZD"

# ═══════════════════════════════════════════════════════
#  REGEX
# ═══════════════════════════════════════════════════════

@mcp.tool()
def regex(pattern: str, text: str) -> str:
    """Test regex pattern against text. Example: regex \\d+ abc123"""
    try:
        matches = re.findall(pattern, text)
        return f"Matches ({len(matches)}):\n" + "\n".join(f"  [{i+1}] {m}" for i, m in enumerate(matches)) if matches else "No matches."
    except re.error as e:
        return f"Invalid regex: {e}"

@mcp.tool()
def regex_groups(pattern: str, text: str) -> str:
    """Test regex with capture groups."""
    try:
        matches = list(re.finditer(pattern, text))
        if not matches: return "No matches."
        lines = []
        for i, m in enumerate(matches):
            lines.append(f"  [{i+1}] {m.group()}")
            for j, g in enumerate(m.groups()):
                lines.append(f"      Group {j+1}: {g}")
        return "\n".join(lines)
    except re.error as e:
        return f"Invalid regex: {e}"

# ═══════════════════════════════════════════════════════
#  TEXT STATS
# ═══════════════════════════════════════════════════════

@mcp.tool()
def text_stats(text: str) -> str:
    """Word count, characters, lines, reading time."""
    words = text.split()
    return f"""Words:     {len(words)}
Chars:     {len(text)}
Lines:     {text.count(chr(10)) + 1}
Sentences: {len(re.findall(r'[.!?]+', text))}
Reading:   ~{max(1, len(words) // 200)} min"""

# ═══════════════════════════════════════════════════════
#  DIFF
# ═══════════════════════════════════════════════════════

@mcp.tool()
def diff(text1: str, text2: str) -> str:
    """Compare two texts. Shows unified diff."""
    d = list(difflib.unified_diff(text1.splitlines(keepends=True), text2.splitlines(keepends=True), lineterm=""))
    return "".join(d) if d else "Identical."

# ═══════════════════════════════════════════════════════
#  PASSWORD STRENGTH
# ═══════════════════════════════════════════════════════

@mcp.tool()
def password_strength(password: str) -> str:
    """Check password strength. Returns score + improvement tips."""
    score = 0
    issues = []
    if len(password) >= 8: score += 1
    else: issues.append("Use >= 8 characters")
    if len(password) >= 12: score += 1
    if re.search(r'[A-Z]', password): score += 1
    else: issues.append("Add uppercase")
    if re.search(r'[a-z]', password): score += 1
    else: issues.append("Add lowercase")
    if re.search(r'\d', password): score += 1
    else: issues.append("Add numbers")
    if re.search(r'[!@#$%^&*()]', password): score += 1
    else: issues.append("Add special chars")
    levels = ["Very Weak", "Weak", "Fair", "Good", "Strong", "Very Strong"]
    bar = "█" * score + "░" * (6 - score)
    result = f"Strength: {levels[min(score, 5)]}\nScore:    [{bar}] {score}/6"
    if issues: result += "\n\nTips:\n" + "\n".join(f"  - {i}" for i in issues)
    return result

# ═══════════════════════════════════════════════════════
#  MARKDOWN TO HTML
# ═══════════════════════════════════════════════════════

@mcp.tool()
def markdown_to_html(text: str) -> str:
    """Convert markdown to HTML (headings, bold, italic, code, lists)."""
    text = re.sub(r'^### (.+)$', r'<h3>\1</h3>', text, flags=re.MULTILINE)
    text = re.sub(r'^## (.+)$', r'<h2>\1</h2>', text, flags=re.MULTILINE)
    text = re.sub(r'^# (.+)$', r'<h1>\1</h1>', text, flags=re.MULTILINE)
    text = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', text)
    text = re.sub(r'\*(.+?)\*', r'<em>\1</em>', text)
    text = re.sub(r'`(.+?)`', r'<code>\1</code>', text)
    text = re.sub(r'^- (.+)$', r'<li>\1</li>', text, flags=re.MULTILINE)
    return text

# ═══════════════════════════════════════════════════════
#  JSON
# ═══════════════════════════════════════════════════════

@mcp.tool()
def json_format(text: str) -> str:
    """Format and validate JSON."""
    try:
        return json.dumps(json.loads(text), indent=2)
    except json.JSONDecodeError as e:
        return f"Invalid JSON: {e}"

@mcp.tool()
def json_minify(text: str) -> str:
    """Minify JSON to single line."""
    try:
        return json.dumps(json.loads(text), separators=(",", ":"))
    except json.JSONDecodeError as e:
        return f"Invalid JSON: {e}"

@mcp.tool()
def json_keys(text: str) -> str:
    """Extract all keys from JSON (dot notation)."""
    try:
        data = json.loads(text)
        def walk(obj, prefix=""):
            keys = []
            if isinstance(obj, dict):
                for k, v in obj.items():
                    p = f"{prefix}.{k}" if prefix else k
                    keys.append(p)
                    keys.extend(walk(v, p))
            elif isinstance(obj, list) and obj:
                keys.extend(walk(obj[0], f"{prefix}[0]"))
            return keys
        return "\n".join(walk(data))
    except json.JSONDecodeError as e:
        return f"Invalid JSON: {e}"

# ═══════════════════════════════════════════════════════
#  IP / NETWORK
# ═══════════════════════════════════════════════════════

@mcp.tool()
def ip_info(ip: str = "") -> str:
    """Get info about an IP. Leave empty for your own IP."""
    url = f"https://ip-api.com/json/{ip}" if ip else "https://ip-api.com/json/"
    data = httpx.get(url).json()
    if data.get("status") == "success":
        return (f"IP: {data['query']}\nCity: {data.get('city','N/A')}\n"
                f"Region: {data.get('regionName','N/A')}\nCountry: {data.get('country','N/A')}\n"
                f"ISP: {data.get('isp','N/A')}\nLat/Lon: {data.get('lat','N/A')}, {data.get('lon','N/A')}")
    return "Failed to get IP info."

@mcp.tool()
def public_ip() -> str:
    """Get your public IP address."""
    return f"Your public IP: {httpx.get('https://api.ipify.org').text}"

# ═══════════════════════════════════════════════════════
#  WIKIPEDIA
# ═══════════════════════════════════════════════════════

@mcp.resource("greeting://{name}")
def get_greeting(name: str) -> str:
    return f"Hello, {name}!"

@mcp.tool()
def wiki_search(query: str, limit: int = 5) -> str:
    """Search Wikipedia for articles."""
    params = {"action": "query", "list": "search", "srsearch": query, "srlimit": limit, "format": "json"}
    results = httpx.get("https://en.wikipedia.org/w/api.php", params=params, headers=HEADERS).json()
    titles = [r["title"] for r in results.get("query", {}).get("search", [])]
    return "\n".join(titles) if titles else "No results."

@mcp.tool()
def wiki_get_page(title: str) -> str:
    """Get summary of a Wikipedia article."""
    resp = httpx.get(f"{WIKI_API}/page/summary/{title}", headers=HEADERS)
    if resp.status_code == 200:
        return resp.json().get("extract", "No content.")
    return f"Article '{title}' not found."

# ═══════════════════════════════════════════════════════
#  WEATHER
# ═══════════════════════════════════════════════════════

@mcp.tool()
def weather(city: str) -> str:
    """Get current weather for a city."""
    r = httpx.get(f"https://wttr.in/{city}?format=3", headers=HEADERS)
    return r.text.strip()

@mcp.tool()
def weather_detail(city: str) -> str:
    """Get detailed weather for a city."""
    r = httpx.get(f"https://wttr.in/{city}?format=%l:+%C+%t+%h+%w", headers=HEADERS)
    return r.text.strip()

# ═══════════════════════════════════════════════════════
#  COLOR
# ═══════════════════════════════════════════════════════

@mcp.tool()
def hex_to_rgb(hex_color: str) -> str:
    """Hex to RGB. Example: #ff5733"""
    h = hex_color.lstrip("#")
    return f"R: {int(h[:2],16)}, G: {int(h[2:4],16)}, B: {int(h[4:6],16)}"

@mcp.tool()
def rgb_to_hex(r: int, g: int, b: int) -> str:
    """RGB to hex."""
    return f"#{r:02x}{g:02x}{b:02x}"

@mcp.tool()
def color_palette(hex_color: str) -> str:
    """Generate complementary, analogous, and triadic colors."""
    hex_color = hex_color.lstrip("#")
    r, g, b = int(hex_color[:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)

    def to_hsl(r, g, b):
        r, g, b = r/255, g/255, b/255
        mx, mn = max(r,g,b), min(r,g,b)
        l = (mx + mn) / 2
        if mx == mn: return 0, 0, l
        d = mx - mn
        s = d / (2 - mx - mn) if l > 0.5 else d / (mx + mn)
        if mx == r: h = (g - b) / d + (6 if g < b else 0)
        elif mx == g: h = (b - r) / d + 2
        else: h = (r - g) / d + 4
        return h / 6 * 360, s * 100, l * 100

    def from_hsl(h, s, l):
        h, s, l = h/360, s/100, l/100
        def hue2(p, q, t):
            if t < 0: t += 1
            if t > 1: t -= 1
            if t < 1/6: return p + (q - p) * 6 * t
            if t < 1/2: return q
            if t < 2/3: return p + (q - p) * (2/3 - t) * 6
            return p
        q = l * (1 + s) if l < 0.5 else l + s - l * s
        p = 2 * l - q
        return (int(hue2(p, q, h/360 + 1/3) * 255),
                int(hue2(p, q, h/360) * 255),
                int(hue2(p, q, h/360 - 1/3) * 255))

    h, s, l = to_hsl(r, g, b)
    def fmt(r, g, b):
        return f"#{r:02x}{g:02x}{b:02x}"
    return (f"Original:        {fmt(r,g,b)}\n"
            f"Complementary:   {fmt(*from_hsl((h+180)%360, s, l))}\n"
            f"Analogous 1:     {fmt(*from_hsl((h+30)%360, s, l))}\n"
            f"Analogous 2:     {fmt(*from_hsl((h-30)%360, s, l))}\n"
            f"Triadic 1:       {fmt(*from_hsl((h+120)%360, s, l))}\n"
            f"Triadic 2:       {fmt(*from_hsl((h+240)%360, s, l))}")

if __name__ == "__main__":
    mcp.run()
